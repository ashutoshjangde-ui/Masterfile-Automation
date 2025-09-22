# app_masterfile.py
import streamlit as st
import json, re, time
from io import BytesIO
from difflib import SequenceMatcher
from openpyxl import load_workbook

# -------------------------- Streamlit Page Setup --------------------------
st.set_page_config(
    page_title="Masterfile Automation",
    page_icon="📦",
    layout="wide"
)

# Subtle UI polish
st.markdown("""
<style>
/* Cards */
.report-card {
  background: #ffffff;
  border: 1px solid #e6e8eb;
  border-radius: 14px;
  padding: 16px 18px;
  box-shadow: 0 1px 2px rgba(0,0,0,0.03);
}

/* Tag chips */
.tag {
  display: inline-block;
  background: #F1F5F9;
  color: #111827;
  border-radius: 999px;
  padding: 2px 10px;
  margin-right: 6px;
  font-size: 12px;
  border: 1px solid #E5E7EB;
}
.ok    { background:#DCFCE7; border-color:#BBF7D0;}
.warn  { background:#FEF3C7; border-color:#FDE68A;}
.err   { background:#FEE2E2; border-color:#FCA5A5;}
</style>
""", unsafe_allow_html=True)

st.title("📦 Masterfile Automation")
st.caption("Map onboarding columns to your masterfile template, preserve formatting, and export in one click.")

# -------------------------- Helpers (same logic as your script) --------------------------
# Fixed assumptions per your last spec:
ONBOARDING_HEADER_ROW = 1
ONBOARDING_DATA_START = 2
MASTER_DISPLAY_ROW    = 1
MASTER_KEYS_ROW       = 2
MASTER_DATA_START     = 3

MASTER_COL_CAP   = 200
ONBOARD_COL_CAP  = 300
EMPTY_STREAK_STOP = 12
ONBOARD_ROW_CAP   = 20000
ROW_EMPTY_STREAK  = 50

def norm(s: str) -> str:
    if s is None: return ""
    x = str(s).strip().lower()
    x = re.sub(r"\s*-\s*en\s*[-_ ]\s*us\s*$", "", x)   # trim " - en-US"
    x = x.replace("–","-").replace("—","-").replace("−","-")
    x = re.sub(r"[._\\-/]+", " ", x)
    x = re.sub(r"[^0-9a-z\\s]+", " ", x)
    return re.sub(r"\\s+", " ", x).strip()

def top_matches(query, candidates, k=3):
    q = norm(query)
    scored = [(SequenceMatcher(None, q, norm(c)).ratio(), c) for c in candidates]
    scored.sort(key=lambda t: t[0], reverse=True)
    return scored[:k]

def read_row_values(ws, row_idx, max_col):
    return [(ws.cell(row=row_idx, column=c).value or "") for c in range(1, max_col+1)]

def detect_used_cols(ws, header_rows, hard_cap, empty_streak_stop):
    max_try = min(ws.max_column, hard_cap)
    last_nonempty, streak = 0, 0
    for c in range(1, max_try+1):
        any_val = False
        for r in header_rows:
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                any_val = True
                break
        if any_val:
            last_nonempty = c
            streak = 0
        else:
            streak += 1
            if streak >= empty_streak_stop:
                break
    return max(last_nonempty, 1)

def detect_used_rows(ws, start_row, used_cols, hard_cap, row_empty_streak):
    last, streak = start_row-1, 0
    max_try = min(ws.max_row, hard_cap)
    for r in range(start_row, max_try+1):
        any_val = any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, used_cols+1))
        if any_val:
            last = r
            streak = 0
        else:
            streak += 1
            if streak >= row_empty_streak:
                break
    if last < start_row-1:
        return start_row-1
    return last

def pick_data_sheet(wb, header_rows, col_cap, empty_streak_stop):
    best_ws, best_score = None, -1
    for ws in wb.worksheets:
        used_cols = detect_used_cols(ws, header_rows, col_cap, empty_streak_stop)
        row_score = 0
        for r in header_rows:
            row_vals = read_row_values(ws, r, used_cols)
            row_score += sum(1 for v in row_vals if str(v).strip())
        if row_score > best_score:
            best_score = row_score
            best_ws = ws
    return best_ws

# -------------------------- Sidebar / Uploads --------------------------
with st.sidebar:
    st.header("⚙️ Inputs")
    onboarding_file = st.file_uploader("Onboarding (.xlsx)", type=["xlsx"])
    masterfile_file = st.file_uploader("Masterfile Template (.xlsx)", type=["xlsx"])
    mapping_file    = st.file_uploader("Mapping (.json)", type=["json"])

    st.markdown("---")
    advanced = st.checkbox("Advanced sheet selection", value=False)
    master_sheet_name = st.text_input("Master sheet name (optional)") if advanced else None
    onboarding_sheet_name = st.text_input("Onboarding sheet name (optional)") if advanced else None

    go = st.button("🚀 Generate Masterfile", use_container_width=True)

# -------------------------- Action --------------------------
if go:
    if not (onboarding_file and masterfile_file and mapping_file):
        st.error("Please upload all three files: Onboarding, Masterfile template, and Mapping JSON.")
        st.stop()

    try:
        mapping_json = json.load(mapping_file)
    except Exception as e:
        st.error(f"Mapping JSON could not be parsed. Error: {e}")
        st.stop()

    # Normalize mapping keys for robust lookup
    MAPPING = {}
    for k, v in mapping_json.items():
        MAPPING[norm(k)] = v[:] if isinstance(v, list) else [v]

    with st.spinner("Reading workbooks…"):
        # Read bytes then to BytesIO for openpyxl
        onboarding_bytes = onboarding_file.read()
        masterfile_bytes = masterfile_file.read()

        # Load workbooks
        mw = load_workbook(BytesIO(masterfile_bytes), keep_links=False)
        ow = load_workbook(BytesIO(onboarding_bytes), data_only=True, read_only=True, keep_links=False)

        # pick sheets
        if master_sheet_name and master_sheet_name in mw.sheetnames:
            ms = mw[master_sheet_name]
        else:
            ms = pick_data_sheet(mw, [MASTER_DISPLAY_ROW, MASTER_KEYS_ROW], MASTER_COL_CAP, EMPTY_STREAK_STOP)

        if onboarding_sheet_name and onboarding_sheet_name in ow.sheetnames:
            osheet = ow[onboarding_sheet_name]
        else:
            osheet = pick_data_sheet(ow, [ONBOARDING_HEADER_ROW], ONBOARD_COL_CAP, EMPTY_STREAK_STOP)

    # Show selected sheets
    colA, colB, colC = st.columns([1.2,1.2,2])
    with colA:
        st.markdown(f"<div class='report-card'><b>📄 Master sheet</b><br><span class='tag ok'>{ms.title}</span></div>", unsafe_allow_html=True)
    with colB:
        st.markdown(f"<div class='report-card'><b>📥 Onboarding sheet</b><br><span class='tag ok'>{osheet.title}</span></div>", unsafe_allow_html=True)

    # Detect widths
    m_used_cols  = detect_used_cols(ms,     [MASTER_DISPLAY_ROW, MASTER_KEYS_ROW], MASTER_COL_CAP,  EMPTY_STREAK_STOP)
    on_used_cols = detect_used_cols(osheet, [ONBOARDING_HEADER_ROW],             ONBOARD_COL_CAP, EMPTY_STREAK_STOP)

    # Read header rows
    master_row1 = read_row_values(ms, MASTER_DISPLAY_ROW, m_used_cols)  # display
    master_row2 = read_row_values(ms, MASTER_KEYS_ROW,    m_used_cols)  # keys
    on_headers  = read_row_values(osheet, ONBOARDING_HEADER_ROW, on_used_cols)  # onboarding headers row 1

    # Build onboarding alias -> column index (only row1 per your rule)
    alias_to_col = {}
    for c in range(1, on_used_cols+1):
        h = osheet.cell(row=ONBOARDING_HEADER_ROW, column=c).value
        if h:
            alias_to_col[norm(h)] = c

    all_onboard_aliases = [h for h in on_headers if str(h).strip()]

    # Map master columns to onboarding columns
    col_map = {}
    unmatched = []
    mapping_report_lines = []

    for mcol in range(1, m_used_cols+1):
        m_disp = master_row1[mcol-1] or ""
        m_key  = master_row2[mcol-1] or ""

        aliases = []
        aliases += MAPPING.get(norm(m_disp), [])
        aliases += MAPPING.get(norm(m_key),  [])
        if m_disp: aliases.append(m_disp)
        if m_key:  aliases.append(m_key)

        resolved = None
        for a in aliases:
            a_norm = norm(a)
            if a_norm in alias_to_col:
                resolved = alias_to_col[a_norm]
                break

        if resolved:
            col_map[mcol] = resolved
            mapping_report_lines.append(f"✅ {m_disp} | {m_key}  →  {osheet.cell(row=ONBOARDING_HEADER_ROW, column=resolved).value}")
        else:
            if norm(m_disp) == norm("Listing Action (List or Unlist)"):
                col_map[mcol] = None  # fill "List"
                mapping_report_lines.append(f"• {m_disp} | {m_key}  →  will fill 'List'")
            else:
                mapping_report_lines.append(f"⚠️ No match: {m_disp} | {m_key}")
                # show suggestions
                for sc, s in top_matches(m_disp, all_onboard_aliases, 3):
                    mapping_report_lines.append(f"   ↳ near display: {s} ({round(sc*100,1)}%)")
                for sc, s in top_matches(m_key, all_onboard_aliases, 3):
                    mapping_report_lines.append(f"   ↳ near key    : {s} ({round(sc*100,1)}%)")
                unmatched.append((m_disp, m_key))

    # Compute onboarding data range (Row 2..end)
    start_on = ONBOARDING_DATA_START
    last_on  = detect_used_rows(osheet, start_on, on_used_cols, ONBOARD_ROW_CAP, ROW_EMPTY_STREAK)

    # Copy rows into template
    rows_copied = 0
    with st.spinner("Copying rows into the template…"):
        for r in range(start_on, last_on+1):
            if not any(osheet.cell(row=r, column=c).value not in (None, "") for c in range(1, on_used_cols+1)):
                continue
            target_row = MASTER_DATA_START + rows_copied
            for mcol in range(1, m_used_cols+1):
                src_col = col_map.get(mcol, "MISSING")
                m_disp = master_row1[mcol-1] or ""
                if src_col is None and norm(m_disp) == norm("Listing Action (List or Unlist)"):
                    ms.cell(row=target_row, column=mcol).value = "List"
                elif isinstance(src_col, int):
                    ms.cell(row=target_row, column=mcol).value = osheet.cell(row=r, column=src_col).value
                # else: leave blank
            rows_copied += 1

    # Save to memory and offer download
    out_bytes = BytesIO()
    mw.save(out_bytes)
    out_bytes.seek(0)

    # Summary cards
    with colC:
        filled = sum(1 for v in col_map.values() if isinstance(v, int))
        total  = len(col_map)
        st.markdown(
            f"""
            <div class='report-card'>
              <b>📊 Summary</b><br>
              <span class='tag ok'>Mapped: {filled}</span>
              <span class='tag warn'>Unmapped: {total-filled}</span>
              <span class='tag'>Rows copied: {rows_copied}</span>
            </div>
            """, unsafe_allow_html=True
        )

    st.success("✅ Masterfile generated!")

    st.download_button(
        "⬇️ Download Final Masterfile",
        data=out_bytes.getvalue(),
        file_name="final_masterfile_real.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    with st.expander("🧭 Mapping report (details)"):
        st.code("\n".join(mapping_report_lines), language="text")

    if unmatched:
        st.warning("Some master columns did not find a match. Open the report above and add the shown synonyms to your mapping JSON if needed.")
